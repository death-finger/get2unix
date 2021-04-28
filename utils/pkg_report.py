from openpyxl.styles import Alignment
from openpyxl import Workbook
import json
import sys
import os


USAGE_MSG = """
Usage:  ./pkg_report <output_excel_file> <ansible_result_files>
Current supported OS release:
    - apt: Ubuntu-18 / Ubuntu-20
    - yum: CentOS-7 / RedHat-7
    - dnf: CentOS-8 / RedHat-8
<ansible_result_files> should be file path joined by "," like:
    ./pkg_report pkg_analyze.xlsx /home/user/RedHat-8.output,/home/user/RedHat-7.output,/home/user/CentOS-7.output
"""


def scan_apt(filename, os_version, work_book):
    raw_data = []
    process_type = 0
    WS = work_book.create_sheet(title="%s" % os_version)
    WS.append(('HOSTNAME', 'NEW_PKGS', 'UPGRADE_PKGS'))

    with open(filename, 'r', encoding='utf8') as f:
        for line in f.readlines():
            raw_data.append(line)

    for line in raw_data:
        data = json.loads(line.replace("'", '"').replace("True", '"True"').replace('False', '"False"'))
        for k, v in data.items():
            new_pkgs = ""
            upgrade_pkgs = ""
            for item in v['stdout_lines']:
                if process_type == 1:
                    new_pkgs = item.lstrip().replace('{a} ', '\n')
                    process_type = 0
                elif process_type == 2:
                    upgrade_pkgs = item.lstrip().rstrip().replace(' ', "\n")
                    process_type = 0
                else:
                    if item == 'The following packages will be upgraded:':
                        process_type = 2
                    elif item == 'The following NEW packages will be installed:':
                        process_type = 1
                    else:
                        process_type = 0
            if (not new_pkgs) and (not upgrade_pkgs):
                new_pkgs = "All the packages are latest"
                upgrade_pkgs = "All the packages are latest"
            WS.append((k, new_pkgs, upgrade_pkgs))
        WS.column_dimensions["A"].width = 20
        WS.column_dimensions["B"].width = 60
        WS.column_dimensions["C"].width = 60
        for row in WS.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)


def scan_yum(filename, os_version, work_book):
    raw_data = []
    WS = work_book.create_sheet(title="%s" % os_version)
    WS.append(('HOSTNAME', 'NEW_PKGS', 'UPGRADE_PKGS'))

    with open(filename, 'r', encoding='utf8') as f:
        for line in f.readlines():
            raw_data.append(line)

    for line in raw_data:
        data = json.loads(line.replace("'", '"').replace("True", '"True"').replace('False', '"False"'))
        for k, v in data.items():
            pkgs_new = ""
            pkgs_update = ""
            if v['changed'] == 'False':
                pkgs_new = "All the packages are latest"
                pkgs_update = "All the packages are latest"
            else:
                for i in v['changes']['updated']:
                    pkgs_update += "%s-%s\n" % (i[0], i[1].split(" from ")[0])
                for i in v['changes']['installed']:
                    pkgs_new += "%s-%s\n" % (i[0], i[1].split(" from ")[0])
            WS.append((k, pkgs_new, pkgs_update))
        WS.column_dimensions["A"].width = 20
        WS.column_dimensions["B"].width = 60
        WS.column_dimensions["C"].width = 60
        for row in WS.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)


def scan_dnf(filename, os_version, work_book):
    raw_data = []
    WS = work_book.create_sheet(title="%s" % os_version)
    WS.append(('HOSTNAME', 'INSTALLED_PKGS', 'REMOVED_PKGS'))

    with open(filename, 'r', encoding='utf8') as f:
        for line in f.readlines():
            raw_data.append(line)

    for line in raw_data:
        data = json.loads(line.replace("'", '"').replace("True", '"True"').replace('False', '"False"'))
        for k, v in data.items():
            pkg_install = ""
            pkg_remove = ""
            if v['changed'] == 'False':
                pkg_install = "All the packages are latest"
                pkg_remove = "All the packages are latest"
            else:
                for i in v['results']:
                    if i.split(" ")[0] == "Installed:":
                        pkg_install += (i.split(" ")[1] + "\n")
                    elif i.split(" ")[0] == "Removed:":
                        pkg_remove += (i.split(" ")[1] + "\n")
            WS.append((k, pkg_install, pkg_remove))
        WS.column_dimensions["A"].width = 20
        WS.column_dimensions["B"].width = 60
        WS.column_dimensions["C"].width = 60
        for row in WS.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)


def main(report_file, output_files):
    work_book = Workbook()
    for item in sorted(output_files):
        os_release = os.path.split(item)[1].split(".")[0]
        if (os_release.split("-")[0] == "Ubuntu") and (os_release.split("-")[1] in ("18", "20")):
            scan_apt(item, os_release, work_book)
        elif (os_release.split("-")[0] in ("RedHat", "CentOS")) and (os_release.split("-")[1] == "7"):
            scan_yum(item, os_release, work_book)
        elif (os_release.split("-")[0] in ("RedHat", "CentOS")) and (os_release.split("-")[1] == "8"):
            scan_dnf(item, os_release, work_book)
    work_book.remove(work_book["Sheet"])
    work_book.save(report_file)


if __name__ == '__main__':
    try:
        REPORT_FILE = sys.argv[1]
        OUTPUT_FILES = sys.argv[2].split(',')
    except:
        print(USAGE_MSG)
    else:
        main(REPORT_FILE, OUTPUT_FILES)
